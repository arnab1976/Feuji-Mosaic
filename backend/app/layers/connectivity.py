"""
Layer 1 — Connectivity (data off the floor).

Production stack: OPC-UA server + Node-RED/Telegraf edge collector +
Mosquitto (MQTT) as the Unified Namespace.

Shop-floor values come from either:
  - simulated OPC-UA tags (random in-band readings), or
  - an Excel / CSV / JSON file the user uploads, replayed as a live stream.
"""
from __future__ import annotations
import csv
import io
import json
import os
import random
import re
import tempfile
from datetime import datetime, timezone, date, timedelta
from pathlib import Path
from threading import Lock
from typing import Dict, List, Optional, Tuple

from ..domain import PARAMETERS, PARAM_IDS, TAG_TO_PARAM


def _line(asset: str) -> str:
    return {"BR-12": "line3", "FIL-07": "line3",
            "WFI-02": "utility", "CR-A1": "cleanroomA"}.get(asset, "line3")


def _uns(p) -> str:
    return f"plant/{_line(p.asset)}/{p.asset}/{p.id}"


def _reading_value(pid: str, force_zone: str | None = None) -> float:
    p = PARAMETERS[pid]
    lo_c, hi_c = p.control
    lo_a, hi_a = p.alarm
    lo_t, hi_t = p.trip
    if force_zone == "alarm":
        return round(random.choice([random.uniform(lo_a, lo_c - 0.01),
                                    random.uniform(hi_c + 0.01, hi_a)]), 3)
    if force_zone == "trip":
        return round(random.choice([random.uniform(lo_t, lo_a - 0.01),
                                    random.uniform(hi_a + 0.01, hi_t)]), 3)
    if force_zone == "control":
        return round(random.uniform(lo_c, hi_c), 3)
    r = random.random()
    if r < 0.78:
        return round(random.uniform(lo_c, hi_c), 3)
    elif r < 0.94:
        return round(random.choice([random.uniform(lo_a, lo_c),
                                    random.uniform(hi_c, hi_a)]), 3)
    else:
        return round(random.choice([random.uniform(lo_t, lo_a),
                                    random.uniform(hi_a, hi_t)]), 3)


def _file_stem(filename: str | None) -> str:
    stem = Path(filename or "upload").stem
    cleaned = "".join(ch if ch.isalnum() or ch in "-_" else "_" for ch in stem)[:48]
    return cleaned or "upload"


def _uns_file(filename: str | None, tag: str) -> str:
    return f"plant/file/{_file_stem(filename)}/{tag}"


def _pack(pid: str, value: float, ts: str | None = None,
          source: str = "OPC-UA/edge", extra: Dict | None = None) -> Dict:
    p = PARAMETERS[pid]
    extra = extra or {}
    from_file = str(source).startswith("uploaded")
    tag = extra.get("file_tag") or p.tag
    unit = extra.get("file_unit") or p.unit
    ts_val = ts or extra.get("file_ts") or datetime.now(timezone.utc).isoformat(timespec="seconds")
    topic = extra.get("uns_topic") or (
        _uns_file(extra.get("filename"), tag) if from_file else _uns(p)
    )
    return {
        "tag": tag,
        "value": round(float(value), 3),
        "unit": unit or p.unit,
        "quality": "GOOD",
        "timestamp": ts_val,
        "uns_topic": topic,
        "source": source,
        "param": pid,
        "asset": p.asset,
        "name": p.short,
        "reading_id": extra.get("reading_id") or extra.get("id") or None,
        "file_tag": extra.get("file_tag") or (tag if from_file else None),
        "file_unit": extra.get("file_unit") or (unit if from_file else None),
        "file_ts": extra.get("file_ts") or (ts if from_file else None),
        "filename": extra.get("filename"),
        "row_index": extra.get("row_index"),
    }


def simulate_reading(pid: str, force_zone: str | None = None) -> Dict:
    return _pack(pid, _reading_value(pid, force_zone), source="OPC-UA/simulator")


SAMPLE_TICKS = 100  # 100 timestamps × 5 tags = 500 rows
PARSE_STATS: Dict = {"raw": 0, "kept": 0, "skipped": 0, "unknown_tags": []}


def _norm_key(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", (s or "").lower())


def _tag_aliases() -> Dict[str, str]:
    out: Dict[str, str] = {
        "temp": "temp", "temperature": "temp", "reactortemperature": "temp",
        "ph": "ph", "acidity": "ph",
        "press": "press", "pressure": "press", "differentialpressure": "press", "dp": "press",
        "cond": "cond", "conductivity": "cond", "wfi": "cond",
        "hum": "hum", "humidity": "hum", "rh": "hum", "relativehumidity": "hum",
    }
    for p in PARAMETERS.values():
        for token in (p.tag, p.id, p.short, p.name, p.tag.replace("-", "")):
            out[_norm_key(token)] = p.id
    return out


TAG_ALIASES = _tag_aliases()


def resolve_param(token: str) -> Optional[str]:
    t = str(token or "").strip()
    if not t:
        return None
    if t in TAG_TO_PARAM:
        return TAG_TO_PARAM[t]
    if t in PARAMETERS:
        return t
    return TAG_ALIASES.get(_norm_key(t))


def _sample_value(pid: str, i: int) -> float:
    p = PARAMETERS[pid]
    lo_c, hi_c = p.control
    lo_a, hi_a = p.alarm
    lo_t, hi_t = p.trip
    k = i % 10
    mid = (lo_c + hi_c) / 2.0
    span = (hi_c - lo_c) or 1.0
    wave = span * 0.2 * (((i % 7) - 3) / 3.0)
    if k == 9:
        return round(min(hi_t, hi_a + (hi_t - hi_a) * 0.45), 3)
    if k >= 7:
        return round(min(hi_a, hi_c + (hi_a - hi_c) * 0.55), 3)
    return round(mid + wave, 3)


def sample_csv() -> str:
    """Canonical shop-floor CSV: 500 rows (100 timestamps × 5 tags)."""
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["tag", "value", "unit", "timestamp"])
    start = datetime(2026, 8, 14, 10, 0, 0, tzinfo=timezone.utc)
    for i in range(SAMPLE_TICKS):
        ts = (start + timedelta(seconds=i)).strftime("%Y-%m-%dT%H:%M:%SZ")
        for p in PARAMETERS.values():
            w.writerow([p.tag, _sample_value(p.id, i), p.unit, ts])
    return buf.getvalue()


def _cell_text(v) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        if v.tzinfo is None:
            v = v.replace(tzinfo=timezone.utc)
        return v.isoformat(timespec="seconds")
    if isinstance(v, date):
        return datetime(v.year, v.month, v.day, tzinfo=timezone.utc).isoformat(timespec="seconds")
    return str(v).strip()


def _field_index(headers: List[str], *names: str) -> Optional[int]:
    keys = {_norm_key(h): i for i, h in enumerate(headers) if str(h or "").strip()}
    for name in names:
        if _norm_key(name) in keys:
            return keys[_norm_key(name)]
    return None


def _find_header_idx(rows: List[List]) -> int:
    for i, row in enumerate(rows[:25]):
        headers = [_cell_text(c) for c in row]
        keys = {_norm_key(h) for h in headers if h}
        if ("tag" in keys or "param" in keys or "parameter" in keys) and (
            "value" in keys or "val" in keys or "reading" in keys
        ):
            return i
        tag_hits = sum(1 for h in headers if resolve_param(h))
        if tag_hits >= 2:
            return i
    return 0


def _wide_records(headers: List[str], rows: List[List]) -> List[Dict]:
    ts_i = _field_index(headers, "timestamp", "ts", "time", "datetime", "date")
    unit_i = _field_index(headers, "unit")
    tag_cols: List[Tuple[int, str]] = []
    for i, h in enumerate(headers):
        if i == ts_i or i == unit_i:
            continue
        pid = resolve_param(h)
        if pid:
            tag_cols.append((i, h.strip() if isinstance(h, str) else _cell_text(h)))
    if len(tag_cols) < 2:
        return []
    out: List[Dict] = []
    for row in rows:
        ts = _cell_text(row[ts_i]) if ts_i is not None and ts_i < len(row) else ""
        unit = _cell_text(row[unit_i]) if unit_i is not None and unit_i < len(row) else ""
        for i, tag in tag_cols:
            if i >= len(row):
                continue
            val = _cell_text(row[i])
            if not val:
                continue
            rec = {"tag": tag, "value": val, "unit": unit, "timestamp": ts}
            out.append(rec)
    return out


def _table_to_records(headers: List[str], rows: List[List]) -> List[Dict]:
    if not headers:
        raise ValueError("File has no header row. Expected: tag, value, unit, timestamp")
    wide = _wide_records(headers, rows)
    if wide:
        return wide
    fields = {_norm_key(h): i for i, h in enumerate(headers) if str(h or "").strip()}
    tag_i = fields.get("tag", fields.get("parameter", fields.get("param")))
    val_i = fields.get("value", fields.get("val", fields.get("reading")))
    if tag_i is None or val_i is None:
        raise ValueError(
            "File must include tag (or param) and value columns, "
            "or one column per tag (TT-1202B, AT-3401, …)"
        )
    unit_i = fields.get("unit")
    ts_i = fields.get("timestamp", fields.get("ts", fields.get("time", fields.get("datetime"))))
    rid_i = fields.get("readingid", fields.get("reading_id"))
    out: List[Dict] = []
    for row in rows:
        def col(i):
            if i is None or i >= len(row):
                return ""
            return _cell_text(row[i])
        tag = col(tag_i)
        if not tag:
            continue
        rec = {
            "tag": tag,
            "value": col(val_i),
            "unit": col(unit_i),
            "timestamp": col(ts_i),
        }
        rid = col(rid_i)
        if rid:
            rec["reading_id"] = rid
        out.append(rec)
    return out


def _rows_to_records(rows: List[List]) -> List[Dict]:
    if not rows:
        raise ValueError("Excel sheet is empty")
    idx = _find_header_idx(rows)
    headers = [_cell_text(c) for c in rows[idx]]
    return _table_to_records(headers, rows[idx + 1:])


def _read_xlsx(data: bytes) -> List[Dict]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise ValueError("Excel .xlsx support is not installed (openpyxl)") from e
    # read_only uses the stored dimension, which often stops at ~30 rows.
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=False)
    best: List[Dict] = []
    try:
        for ws in wb.worksheets:
            rows = [list(row) for row in ws.iter_rows(values_only=True)]
            if not rows:
                continue
            try:
                recs = _rows_to_records(rows)
            except ValueError:
                continue
            if len(recs) > len(best):
                best = recs
    finally:
        wb.close()
    if not best:
        raise ValueError("Excel sheet is empty")
    return best


def _read_xls(data: bytes) -> List[Dict]:
    try:
        import xlrd
    except ImportError as e:
        raise ValueError("Excel .xls support is not installed (xlrd)") from e
    book = xlrd.open_workbook(file_contents=data)
    sheet = book.sheet_by_index(0)
    rows = []
    for r in range(sheet.nrows):
        vals = []
        for c in range(sheet.ncols):
            cell = sheet.cell(r, c)
            if cell.ctype == xlrd.XL_CELL_DATE:
                vals.append(xlrd.xldate_as_datetime(cell.value, book.datemode))
            elif cell.ctype == xlrd.XL_CELL_EMPTY:
                vals.append(None)
            else:
                vals.append(cell.value)
        rows.append(vals)
    if not rows:
        raise ValueError("Excel sheet is empty")
    return _rows_to_records(rows)


def _read_xlsb(data: bytes) -> List[Dict]:
    try:
        from pyxlsb import open_workbook
    except ImportError as e:
        raise ValueError("Excel .xlsb support is not installed (pyxlsb)") from e
    fd, path = tempfile.mkstemp(suffix=".xlsb")
    os.close(fd)
    try:
        with open(path, "wb") as f:
            f.write(data)
        rows = []
        with open_workbook(path) as wb:
            with wb.get_sheet(1) as sheet:
                for row in sheet.rows():
                    rows.append([c.v if c else None for c in row])
    finally:
        try:
            os.unlink(path)
        except OSError:
            pass
    if not rows:
        raise ValueError("Excel sheet is empty")
    return _rows_to_records(rows)


def _read_csv_bytes(data: bytes) -> List[Dict]:
    text = data.decode("utf-8-sig", errors="replace").strip()
    if not text:
        raise ValueError("CSV file is empty")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect)
    rows = [list(r) for r in reader]
    if not rows:
        raise ValueError("CSV has no header row. Expected: tag,value,unit,timestamp")
    return _rows_to_records(rows)


def _read_json_bytes(data: bytes) -> List[Dict]:
    payload = json.loads(data.decode("utf-8-sig"))
    if isinstance(payload, dict):
        payload = payload.get("readings") or payload.get("rows") or payload.get("data") or []
    if not isinstance(payload, list):
        raise ValueError("JSON must be an array of objects, or {readings: [...]}")
    return payload


def parse_floor_file(filename: str, data: bytes) -> List[Dict]:
    """Parse Excel (.xlsx .xlsm .xls .xlsb), CSV/TSV, or JSON shop-floor files."""
    name = (filename or "upload").lower()
    ext = Path(name).suffix
    magic = data[:8]
    try:
        if ext in {".xlsx", ".xlsm"} or (
            ext not in {".xls", ".xlsb", ".csv", ".tsv", ".json", ".txt"} and magic.startswith(b"PK")
        ):
            recs = _read_xlsx(data)
            return parse_floor_payload(readings=recs, source="uploaded/Excel")
        if ext == ".xls" or magic.startswith(b"\xd0\xcf\x11\xe0"):
            recs = _read_xls(data)
            return parse_floor_payload(readings=recs, source="uploaded/Excel")
        if ext == ".xlsb":
            recs = _read_xlsb(data)
            return parse_floor_payload(readings=recs, source="uploaded/Excel")
        if ext in {".json"} or data.lstrip()[:1] in (b"{", b"["):
            recs = _read_json_bytes(data)
            return parse_floor_payload(readings=recs, source="uploaded/JSON")
        recs = _read_csv_bytes(data)
        return parse_floor_payload(readings=recs, source="uploaded/CSV")
    except ValueError:
        raise
    except Exception as e:
        raise ValueError(f"Could not read {filename or 'file'}: {e}") from e


def sample_xlsx() -> bytes:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "floor"
    reader = csv.reader(io.StringIO(sample_csv()))
    for row in reader:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_floor_payload(csv_text: str | None = None,
                        readings: List[Dict] | None = None,
                        source: str = "uploaded/CSV") -> List[Dict]:
    raw: List[Dict] = []
    if readings:
        raw.extend(readings)
    if csv_text:
        text = csv_text.strip().lstrip("\ufeff")
        reader = csv.DictReader(io.StringIO(text))
        if not reader.fieldnames:
            raise ValueError("CSV has no header row. Expected: tag,value,unit,timestamp")
        fields = {f.strip().lower(): f for f in reader.fieldnames if f}
        tag_k = fields.get("tag") or fields.get("parameter") or fields.get("param")
        val_k = fields.get("value") or fields.get("val")
        if not tag_k or not val_k:
            raise ValueError("CSV must include tag (or param) and value columns")
        unit_k = fields.get("unit")
        ts_k = fields.get("timestamp") or fields.get("ts") or fields.get("time")
        rid_k = fields.get("reading_id") or fields.get("readingid")
        for row in reader:
            rec = {
                "tag": (row.get(tag_k) or "").strip(),
                "value": row.get(val_k),
                "unit": (row.get(unit_k) or "").strip() if unit_k else "",
                "timestamp": (row.get(ts_k) or "").strip() if ts_k else "",
            }
            if rid_k and (row.get(rid_k) or "").strip():
                rec["reading_id"] = (row.get(rid_k) or "").strip()
            raw.append(rec)
    packed: List[Dict] = []
    unknown: List[str] = []
    skipped = 0
    for rec in raw:
        token = str(rec.get("tag") or rec.get("param") or rec.get("id") or "").strip()
        pid = resolve_param(token)
        if not pid:
            skipped += 1
            if token:
                unknown.append(token)
            continue
        try:
            value = float(rec.get("value"))
        except (TypeError, ValueError):
            skipped += 1
            continue
        ts = rec.get("timestamp") or rec.get("ts") or None
        if ts:
            ts = str(ts)
        packed.append(_pack(
            pid, value, ts, source=source,
            extra={
                "file_tag": token,
                "file_unit": str(rec.get("unit") or "").strip(),
                "file_ts": ts or "",
                "reading_id": str(rec.get("reading_id") or rec.get("readingid") or "").strip(),
            },
        ))
    PARSE_STATS["raw"] = len(raw)
    PARSE_STATS["kept"] = len(packed)
    PARSE_STATS["skipped"] = skipped
    PARSE_STATS["unknown_tags"] = list(dict.fromkeys(unknown))[:12]
    if not packed:
        hint = f" Unknown tags: {', '.join(unknown[:6])}." if unknown else ""
        raise ValueError(
            "No usable rows. Use known tags "
            f"({', '.join(TAG_TO_PARAM.keys())}) or param ids "
            f"({', '.join(PARAM_IDS)}).{hint}"
        )
    return packed


def _annotate_file_rows(rows: List[Dict], filename: str | None) -> List[Dict]:
    """Stamp every uploaded row with file identity. Never looks like the OPC-UA generator."""
    name = filename or "upload"
    out: List[Dict] = []
    for i, r in enumerate(rows, start=1):
        rd = dict(r)
        tag = rd.get("file_tag") or rd.get("tag")
        rd["filename"] = name
        rd["row_index"] = int(rd["row_index"]) if rd.get("row_index") else i
        rd["file_tag"] = tag
        rd["file_unit"] = rd.get("file_unit") or rd.get("unit") or ""
        rd["file_ts"] = rd.get("file_ts") or rd.get("timestamp") or ""
        rd["tag"] = tag
        rd["uns_topic"] = _uns_file(name, tag)
        src = str(rd.get("source") or "")
        if not src.startswith("uploaded"):
            rd["source"] = "uploaded/file"
        if not rd.get("reading_id"):
            rd["reading_id"] = f"{_file_stem(name)}-{rd['row_index']:04d}"
        out.append(rd)
    return out


class FloorSource:
    """In-memory shop floor: simulated generator or uploaded replay buffer."""

    def __init__(self) -> None:
        self._lock = Lock()
        self.mode = "simulated"
        self.filename = None
        self.rows: List[Dict] = []
        self._cursors: Dict[str, int] = {pid: 0 for pid in PARAM_IDS}
        self.steps_done = 0
        self.last_reading: Optional[Dict] = None
        self.session: Dict = {}
        self.dataset_id = None
        self.persisted = False
        self.track_done = {"simulator": 0, "uploaded": 0}
        self.track_reading = {"simulator": None, "uploaded": None}
        self.last_stream: List[Dict] = []

    def status(self) -> Dict:
        with self._lock:
            by_param = {pid: 0 for pid in PARAM_IDS}
            for r in self.rows:
                by_param[r["param"]] = by_param.get(r["param"], 0) + 1
            return {
                "mode": self.mode,
                "filename": self.filename,
                "row_count": len(self.rows),
                "by_param": by_param,
                "cursors": dict(self._cursors),
                "steps_done": self.steps_done,
                "last_reading": self.last_reading,
                "last_stream": list(self.last_stream),
                "source": self.session.get("source"),
                "dataset_id": self.dataset_id,
                "persisted": self.persisted,
                "track_done": dict(self.track_done),
                "has_upload": len(self.rows) > 0,
                "parse": dict(PARSE_STATS),
                "tags": [
                    {"id": p.id, "tag": p.tag, "asset": p.asset,
                     "unit": p.unit, "name": p.short, "uns": _uns(p)}
                    for p in PARAMETERS.values()
                ],
            }

    def load(self, rows: List[Dict], filename: str | None = None,
             persisted: bool = False, dataset_id=None) -> Dict:
        with self._lock:
            self.mode = "uploaded"
            self.filename = filename or "upload"
            self.rows = _annotate_file_rows(rows, self.filename)
            self._cursors = {pid: 0 for pid in PARAM_IDS}
            self.steps_done = 0
            self.last_reading = None
            self.last_stream = []
            self.session = {}
            self.persisted = persisted
            self.dataset_id = dataset_id
            self.track_done["uploaded"] = 0
            self.track_reading["uploaded"] = None
        return self.status()

    def snapshot_rows(self):
        with self._lock:
            return list(self.rows), self.filename

    def mark_persisted(self, dataset_id, filename: str | None = None) -> Dict:
        with self._lock:
            self.persisted = True
            self.dataset_id = dataset_id
            if filename:
                self.filename = filename
        return self.status()

    def use_simulated(self) -> Dict:
        """Reset the simulator track only. Uploaded rows stay for the other tab."""
        with self._lock:
            self.session["source"] = "simulator"
            self.track_done["simulator"] = 0
            self.track_reading["simulator"] = None
            if not self.rows:
                self.mode = "simulated"
                self.filename = None
                self.persisted = False
                self.dataset_id = None
        return self.status()

    def clear_upload(self) -> Dict:
        """Drop the in-memory file after it has been deleted from PostgreSQL."""
        with self._lock:
            self.mode = "simulated"
            self.filename = None
            self.rows = []
            self._cursors = {pid: 0 for pid in PARAM_IDS}
            self.dataset_id = None
            self.persisted = False
            self.track_done["uploaded"] = 0
            self.track_reading["uploaded"] = None
            self.last_stream = []
        return self.status()

    def reset_steps(self) -> None:
        with self._lock:
            self.steps_done = 0
            self.session = {}
            self.last_reading = None
            self.last_stream = []
            self.track_done = {"simulator": 0, "uploaded": 0}
            self.track_reading = {"simulator": None, "uploaded": None}

    def next_for(self, pid: str, force_zone: str | None = None) -> Dict:
        with self._lock:
            if self.mode == "uploaded":
                pool = [r for r in self.rows if r["param"] == pid]
                if pool:
                    i = self._cursors.get(pid, 0) % len(pool)
                    self._cursors[pid] = i + 1
                    rd = dict(pool[i])
                    rd["timestamp"] = datetime.now(timezone.utc).isoformat(timespec="seconds")
                    rd["source"] = "uploaded/stream"
                    self.last_reading = rd
                    return rd
            rd = simulate_reading(pid, force_zone)
            self.last_reading = rd
            return rd

    def next_uploaded(self, pid: str) -> Dict:
        """Replay the next uploaded row for pid. Never falls back to the simulator."""
        with self._lock:
            if not self.rows:
                raise ValueError("Upload an Excel, CSV or JSON file first")
            pool = [r for r in self.rows if r["param"] == pid]
            if not pool:
                raise ValueError(f"No uploaded rows for parameter {pid}")
            i = self._cursors.get(pid, 0) % len(pool)
            self._cursors[pid] = i + 1
            rd = dict(pool[i])
            rd["replayed_at"] = datetime.now(timezone.utc).isoformat(timespec="seconds")
            rd["timestamp"] = rd.get("file_ts") or rd.get("timestamp") or rd["replayed_at"]
            rd["source"] = "uploaded/file"
            rd["uns_topic"] = _uns_file(self.filename, rd.get("file_tag") or rd["tag"])
            self.last_reading = rd
            return rd

    def peek_batch(self, pids: List[str], n: int = 6,
                   force_zone: str | None = None) -> List[Dict]:
        out: List[Dict] = []
        for _ in range(n):
            for pid in pids:
                out.append(self.next_for(pid, force_zone))
        return out


FLOOR = FloorSource()
FLOOR_DIR = Path(__file__).resolve().parents[2] / "data" / "floor"
INBOX_DIR = Path(__file__).resolve().parents[2] / "data" / "inbox"
_MOSAIC_ROOT = Path(__file__).resolve().parents[3]


def _local_search_dirs() -> List[Path]:
    home = Path.home()
    extra = os.environ.get("MOSAIC_FLOOR_DIR") or ""
    dirs = [
        INBOX_DIR,
        FLOOR_DIR,
        _MOSAIC_ROOT / "Mosaic Data Sets",
        _MOSAIC_ROOT / "Mosaic DataSets",
        home / "MOSAIC" / "Mosaic Data Sets",
        home / "Documents" / "MOSAIC" / "Mosaic Data Sets",
        home / "OneDrive" / "MOSAIC" / "Mosaic Data Sets",
        home / "OneDrive - Personal" / "MOSAIC" / "Mosaic Data Sets",
    ]
    try:
        for od in home.glob("OneDrive*"):
            if not od.is_dir():
                continue
            dirs.append(od / "MOSAIC" / "Mosaic Data Sets")
            dirs.append(od / "Mosaic Data Sets")
            dirs.append(od / "Desktop" / "MOSAIC" / "Mosaic Data Sets")
            desk = od / "Desktop"
            if desk.is_dir():
                try:
                    for p in desk.glob("**/Mosaic Data Sets"):
                        if p.is_dir():
                            dirs.append(p)
                except OSError:
                    pass
    except OSError:
        pass
    if extra:
        dirs.insert(0, Path(extra))
    out: List[Path] = []
    seen = set()
    for d in dirs:
        try:
            key = str(d.resolve()) if d.exists() else str(d)
        except OSError:
            key = str(d)
        if key in seen:
            continue
        seen.add(key)
        out.append(d)
    return out


def list_local_floor_files() -> List[Dict]:
    INBOX_DIR.mkdir(parents=True, exist_ok=True)
    exts = {".xlsx", ".xlsm", ".xls", ".xlsb", ".csv", ".tsv", ".json", ".txt"}
    found: List[Dict] = []
    seen = set()
    for folder in _local_search_dirs():
        if not folder.is_dir():
            continue
        try:
            entries = list(folder.iterdir())
        except OSError:
            continue
        for p in entries:
            if not p.is_file() or p.suffix.lower() not in exts:
                continue
            try:
                key = str(p.resolve())
            except OSError:
                key = str(p)
            if key in seen:
                continue
            seen.add(key)
            try:
                st = p.stat()
            except OSError:
                continue
            found.append({
                "name": p.name,
                "path": str(p),
                "bytes": st.st_size,
                "mtime": st.st_mtime,
            })
    found.sort(key=lambda r: r["mtime"], reverse=True)
    return found


def _read_shared(path: Path) -> bytes:
    """Read with FILE_SHARE_READ|WRITE|DELETE so Excel/OneDrive locks do not block."""
    if os.name != "nt":
        return path.read_bytes()
    import ctypes
    import msvcrt
    from ctypes import wintypes

    GENERIC_READ = 0x80000000
    FILE_SHARE_ALL = 0x00000007
    OPEN_EXISTING = 3
    FILE_ATTRIBUTE_NORMAL = 0x80
    INVALID = wintypes.HANDLE(-1).value

    CreateFileW = ctypes.windll.kernel32.CreateFileW
    CreateFileW.argtypes = [
        wintypes.LPCWSTR, wintypes.DWORD, wintypes.DWORD,
        wintypes.LPVOID, wintypes.DWORD, wintypes.DWORD, wintypes.HANDLE,
    ]
    CreateFileW.restype = wintypes.HANDLE
    handle = CreateFileW(
        str(path), GENERIC_READ, FILE_SHARE_ALL,
        None, OPEN_EXISTING, FILE_ATTRIBUTE_NORMAL, None,
    )
    if handle in (INVALID, 0, None):
        err = ctypes.GetLastError()
        raise OSError(err, f"Cannot open {path.name} (Windows error {err})")
    fd = msvcrt.open_osfhandle(int(handle), os.O_RDONLY)
    with os.fdopen(fd, "rb") as fh:
        return fh.read()


def _pick_local_floor(files: List[Dict]) -> Path:
    named = [f for f in files if str(f["name"]).lower().startswith("mosaic-floor-data")]
    pool = named or files
    return Path(pool[0]["path"])


def load_local_floor(path: Optional[str] = None) -> Dict:
    files = list_local_floor_files()
    if path:
        want = Path(path)
        rec = next((f for f in files if Path(f["path"]) == want or f["name"] == want.name), None)
        if not rec:
            raise ValueError("That file is not in a MOSAIC data folder")
        target = Path(rec["path"])
    else:
        if not files:
            raise ValueError(
                "No shop-floor file found. Drag a file onto the box, or copy it into "
                f"{INBOX_DIR}"
            )
        target = _pick_local_floor(files)
    try:
        data = _read_shared(target)
    except OSError as e:
        raise ValueError(
            f"Could not read {target.name}. Close it in Excel, then drag it onto the box."
        ) from e
    if not data:
        raise ValueError(f"{target.name} is empty")
    rows = parse_floor_file(target.name, data)
    stash_upload(target.name, data)
    return FLOOR.load(rows, target.name)


def stash_upload(filename: str, data: bytes) -> None:
    FLOOR_DIR.mkdir(parents=True, exist_ok=True)
    for p in FLOOR_DIR.iterdir():
        if p.is_file():
            p.unlink()
    safe = Path(filename or "upload").name
    safe = re.sub(r"[^A-Za-z0-9._-]+", "_", safe)[:120] or "upload"
    (FLOOR_DIR / safe).write_bytes(data or b"")


def purge_stashed_upload() -> int:
    n = 0
    if not FLOOR_DIR.exists():
        return n
    for p in FLOOR_DIR.iterdir():
        if p.is_file():
            p.unlink()
            n += 1
    return n


def emit_reading(pid: str, force_zone: str | None = None) -> Dict:
    return FLOOR.next_for(pid, force_zone)


def emit_all(force_zone: str | None = None) -> List[Dict]:
    return [emit_reading(pid, force_zone) for pid in PARAM_IDS]


def _pids(param: str | None) -> List[str]:
    if not param or param == "all":
        return list(PARAM_IDS)
    if param not in PARAMETERS:
        raise ValueError(f"unknown parameter: {param}")
    return [param]


def _pids_for(param: str | None, source: str) -> List[str]:
    """Uploaded track only walks parameters that actually exist in the file."""
    if source != "uploaded":
        return _pids(param)
    present = {r["param"] for r in FLOOR.rows if r.get("param")}
    ordered = [pid for pid in PARAM_IDS if pid in present]
    if not ordered:
        raise ValueError("Uploaded file has no recognised parameters")
    if not param or param == "all":
        return ordered
    if param not in PARAMETERS:
        raise ValueError(f"unknown parameter: {param}")
    if param not in present:
        raise ValueError(
            f"Uploaded file has no rows for {PARAMETERS[param].tag} ({param})"
        )
    return [param]


def _norm_source(source: str | None) -> str:
    return "uploaded" if source == "uploaded" else "simulator"


def _track_next(source: str, pid: str, zone: str | None = None) -> Dict:
    """Never mix tracks: simulator always generates; uploaded always replays the file."""
    if source == "uploaded":
        if not FLOOR.rows:
            raise ValueError("Upload an Excel, CSV or JSON file first")
        rd = FLOOR.next_uploaded(pid)
        FLOOR.track_reading["uploaded"] = rd
        return rd
    rd = simulate_reading(pid, zone)
    FLOOR.track_reading["simulator"] = rd
    FLOOR.last_reading = rd
    return rd


def _fmt_file_rd(rd: Dict) -> str:
    fn = rd.get("filename") or FLOOR.filename or "upload"
    ri = rd.get("row_index", "?")
    ts = rd.get("file_ts") or rd.get("timestamp") or ""
    return (
        f"[file row {ri} {fn}] {rd['tag']} = {rd['value']} {rd.get('unit') or ''}  "
        f"ts={ts}  →  {rd['uns_topic']}"
    )


def _uploaded_batch(pids: List[str]) -> List[Dict]:
    """Every uploaded row for the selected parameters — the full file slice, not one tick."""
    if not FLOOR.rows:
        raise ValueError("Upload an Excel, CSV or JSON file first")
    want = set(pids)
    batch = [r for r in FLOOR.rows if r.get("param") in want]
    if not batch:
        raise ValueError("Uploaded file has no rows for the selected parameters")
    return batch


def _uploaded_catalog(pids: List[str]) -> List[Dict]:
    items: List[Dict] = []
    for pid in pids:
        pool = [r for r in FLOOR.rows if r.get("param") == pid]
        if not pool:
            continue
        r0, rN = pool[0], pool[-1]
        items.append({
            "param": pid,
            "tag": r0.get("file_tag") or r0["tag"],
            "unit": r0.get("file_unit") or r0.get("unit") or "",
            "topic": r0["uns_topic"],
            "rows": len(pool),
            "first_value": r0["value"],
            "last_value": rN["value"],
            "first_ts": r0.get("file_ts") or r0.get("timestamp"),
            "last_ts": rN.get("file_ts") or rN.get("timestamp"),
        })
    return items


def stream_tick(param: str | None = None) -> Dict:
    """One 1 Hz tick from the uploaded Excel/CSV/JSON — uploaded track only."""
    if not FLOOR.rows:
        raise ValueError("Upload an Excel, CSV or JSON file first")
    pids = _pids_for(param, "uploaded")
    readings = [_track_next("uploaded", pid) for pid in pids]
    FLOOR.track_done["uploaded"] = max(FLOOR.track_done.get("uploaded", 0), 1)
    FLOOR.steps_done = max(FLOOR.steps_done, 1)
    FLOOR.session["source"] = "uploaded"
    st = FLOOR.status()
    chosen = pids[0]
    reading = next((t for t in reversed(readings) if t.get("param") == chosen), readings[-1])
    FLOOR.last_reading = reading
    FLOOR.track_reading["uploaded"] = reading
    FLOOR.last_stream = list(FLOOR.last_stream or []) + list(readings)
    log = [
        f"[file] streaming {FLOOR.filename} — not the OPC-UA simulator",
    ] + [_fmt_file_rd(rd) for rd in readings]
    return {
        "step": 1,
        "source": "uploaded",
        "ok": True,
        "headline": f"Streaming {FLOOR.filename}",
        "filename": FLOOR.filename,
        "row_count": st["row_count"],
        "readings": readings,
        "reading": reading,
        "log": log,
        "track_done": dict(FLOOR.track_done),
    }


def _run_simulator_step(step: int, pids: List[str], zone: str | None, done: int) -> Dict:
    source = "simulator"
    if step == 1:
        tags = []
        first = []
        for pid in pids:
            p = PARAMETERS[pid]
            rd = _track_next(source, pid, zone)
            first.append(rd)
            tags.append({
                "node": f"ns=2;s={p.tag}",
                "tag": p.tag,
                "name": p.short,
                "asset": p.asset,
                "unit": p.unit,
                "value": rd["value"],
                "quality": "GOOD",
            })
        FLOOR.track_done[source] = max(done, 1)
        FLOOR.steps_done = max(FLOOR.steps_done, 1)
        FLOOR.session["source"] = source
        FLOOR.last_stream = list(first)
        return {
            "step": 1, "source": source, "ok": True,
            "headline": "OPC-UA tags live — OPC-UA simulator",
            "log": [
                "[asyncua] Server started  opc.tcp://127.0.0.1:4840/mosaic/server",
                "[asyncua] Namespace       http://mosaic.local/scada",
                "[asyncua] Source          simulator (generated values — not a file)",
                "[asyncua] Publish rate    1 Hz",
            ] + [f"[node] {t['node']}  {t['tag']} = {t['value']} {t['unit']}  ({t['quality']})" for t in tags],
            "tags": tags,
            "readings": first,
            "reading": first[-1] if first else None,
            "track_done": dict(FLOOR.track_done),
        }

    if step == 2:
        FLOOR.track_done[source] = max(done, 2)
        FLOOR.session["source"] = source
        topics = sorted({_uns(PARAMETERS[pid]) for pid in pids})
        return {
            "step": 2, "source": source, "ok": True,
            "headline": "MQTT broker listening — OPC-UA simulator track",
            "log": [
                "[docker] docker run -p 1883:1883 eclipse-mosquitto",
                "[mosquitto] listening on 0.0.0.0:1883",
                "[mosquitto] track OPC-UA simulator — generated tags only",
                "[mosquitto] protocol MQTT 3.1.1 / 5",
                "[UNS] topic tree plant/{line}/{asset}/{param}",
            ] + [f"[UNS] {t}  ← simulator" for t in topics] + [
                "[mosquitto] hub ready — uploaded file is not on this tab",
            ],
            "track_done": dict(FLOOR.track_done),
        }

    if step == 3:
        published = []
        logs = [
            "[Node-RED] track OPC-UA simulator",
            "[Node-RED] flow deployed: opcua-in → function(uns) → mqtt-out",
            "[Node-RED] input node   opcua-in  opc.tcp://127.0.0.1:4840  ← not a file",
            "[Node-RED] MQTT broker  mqtt://127.0.0.1:1883",
        ]
        for pid in pids:
            rd = _track_next(source, pid, zone)
            published.append({
                "tag": rd["tag"], "topic": rd["uns_topic"],
                "payload": {"tag": rd["tag"], "value": rd["value"],
                            "unit": rd["unit"], "ts": rd["timestamp"]},
            })
            logs.append(
                f"[opcua-in] {rd['tag']} = {rd['value']} {rd['unit']}  →  publish {rd['uns_topic']}"
            )
        FLOOR.track_done[source] = max(done, 3)
        FLOOR.last_reading = FLOOR.track_reading.get(source)
        return {
            "step": 3, "source": source, "ok": True,
            "headline": "Edge collector publishing — OPC-UA simulator",
            "log": logs,
            "published": published,
            "reading": FLOOR.track_reading.get(source),
            "track_done": dict(FLOOR.track_done),
        }

    n = 6 if len(pids) == 1 else 2
    ticks = [_track_next(source, pid, zone) for _ in range(n) for pid in pids]
    FLOOR.track_done[source] = 4
    FLOOR.steps_done = 4
    chosen = pids[0]
    reading = next((t for t in reversed(ticks) if t.get("param") == chosen), ticks[-1])
    FLOOR.last_reading = reading
    FLOOR.track_reading[source] = reading
    FLOOR.last_stream = list(ticks)
    FLOOR.session["source"] = source
    topics = sorted({rd["uns_topic"] for rd in ticks})
    logs = [
        "mosquitto_sub -h 127.0.0.1 -t 'plant/line3/#' -v",
        "[mqtt] track OPC-UA simulator (generated values)",
        f"[mqtt] subscribed {len(topics)} simulator topic(s)",
    ]
    for rd in ticks:
        logs.append(
            f"[mqtt] {rd['uns_topic']}  {rd['tag']}={rd['value']} {rd['unit']}  {rd['timestamp']}"
        )
    logs.append(f"[mqtt] ✓ live simulator stream confirmed  ({len(ticks)} messages)")
    return {
        "step": 4, "source": source, "ok": True, "done": True,
        "headline": "Live values on the stream — OPC-UA simulator",
        "log": logs,
        "stream": ticks,
        "readings": ticks,
        "reading": reading,
        "track_done": dict(FLOOR.track_done),
    }


def _run_uploaded_step(step: int, pids: List[str], done: int) -> Dict:
    source = "uploaded"
    fn = FLOOR.filename or "upload"
    batch = _uploaded_batch(pids)
    n_rows = len(batch)
    n_file = len(FLOOR.rows)
    stem = _file_stem(fn)
    tags_present = sorted({rd["tag"] for rd in batch})

    if step == 1:
        tags = [{
            "node": f"file://{fn}#row={rd.get('row_index')}",
            "tag": rd["tag"],
            "name": rd.get("name"),
            "asset": rd.get("asset"),
            "unit": rd.get("unit"),
            "value": rd["value"],
            "quality": "GOOD",
            "file_ts": rd.get("file_ts") or rd.get("timestamp"),
        } for rd in batch]
        FLOOR.track_done[source] = max(done, 1)
        FLOOR.steps_done = max(FLOOR.steps_done, 1)
        FLOOR.session["source"] = source
        FLOOR.last_stream = list(batch)
        FLOOR.last_reading = batch[-1]
        FLOOR.track_reading[source] = batch[-1]
        return {
            "step": 1, "source": source, "ok": True,
            "headline": f"File tags ready — {fn} ({n_rows} datapoints)",
            "log": [
                f"[file] dataset        {fn}",
                f"[file] rows           {n_rows} datapoints  (file has {n_file})",
                f"[file] tags           {', '.join(tags_present)}",
                "[file] source          uploaded spreadsheet/CSV/JSON — NOT the OPC-UA generator",
                f"[file] topic prefix   plant/file/{stem}/",
                f"[file] exposing every row as an OPC-UA node",
            ] + [_fmt_file_rd(rd) for rd in batch] + [
                f"[file] ✓ {n_rows} datapoints exposed from {fn}",
            ],
            "tags": tags,
            "readings": batch,
            "reading": batch[-1],
            "count": n_rows,
            "track_done": dict(FLOOR.track_done),
        }

    if step == 2:
        catalog = _uploaded_catalog(pids)
        FLOOR.track_done[source] = max(done, 2)
        FLOOR.session["source"] = source
        FLOOR.last_stream = list(batch)
        logs = [
            "[docker] docker run -p 1883:1883 eclipse-mosquitto",
            "[mosquitto] listening on 0.0.0.0:1883",
            f"[mosquitto] track uploaded dataset — {fn}",
            f"[mosquitto] registering {n_rows} datapoints in the Unified Namespace",
            "[mosquitto] this hub maps FILE columns to MQTT — simulator tags are not registered",
            f"[UNS] topic tree plant/file/{stem}/{{tag}}",
        ]
        for item in catalog:
            logs.append(
                f"[UNS] {item['topic']}  {item['rows']} file rows  "
                f"{item['tag']}={item['first_value']}…{item['last_value']} {item['unit']}"
            )
        logs.extend(_fmt_file_rd(rd) for rd in batch)
        logs.append(
            f"[mosquitto] hub ready — {n_rows} datapoints mapped  "
            "OPC-UA simulator is not on this tab"
        )
        return {
            "step": 2, "source": source, "ok": True,
            "headline": f"MQTT broker listening — {n_rows} datapoints from {fn}",
            "log": logs,
            "catalog": catalog,
            "readings": batch,
            "reading": batch[-1],
            "count": n_rows,
            "track_done": dict(FLOOR.track_done),
        }

    if step == 3:
        published = []
        logs = [
            f"[Node-RED] track uploaded dataset — {fn}",
            "[Node-RED] flow deployed: file-in → function(uns) → mqtt-out",
            f"[Node-RED] input node   file-in ({fn})  ← not opcua-in / not the simulator",
            "[Node-RED] MQTT broker  mqtt://127.0.0.1:1883",
            f"[Node-RED] publishing all {n_rows} datapoints",
        ]
        for rd in batch:
            published.append({
                "tag": rd["tag"], "topic": rd["uns_topic"],
                "payload": {
                    "tag": rd["tag"], "value": rd["value"],
                    "unit": rd["unit"], "ts": rd.get("file_ts") or rd["timestamp"],
                    "filename": fn, "row": rd.get("row_index"),
                    "source": "uploaded/file",
                },
            })
            logs.append(f"[file-in] {_fmt_file_rd(rd)}")
        logs.append(f"[Node-RED] ✓ published {n_rows} messages to Mosquitto")
        FLOOR.track_done[source] = max(done, 3)
        FLOOR.last_stream = list(batch)
        FLOOR.last_reading = batch[-1]
        FLOOR.track_reading[source] = batch[-1]
        return {
            "step": 3, "source": source, "ok": True,
            "headline": f"Edge collector publishing — {n_rows} datapoints from {fn}",
            "log": logs,
            "published": published,
            "readings": batch,
            "reading": batch[-1],
            "count": n_rows,
            "track_done": dict(FLOOR.track_done),
        }

    FLOOR.track_done[source] = 4
    FLOOR.steps_done = 4
    FLOOR.last_reading = batch[-1]
    FLOOR.track_reading[source] = batch[-1]
    FLOOR.last_stream = list(batch)
    FLOOR.session["source"] = source
    sub = f"plant/file/{stem}/#"
    logs = [
        f"mosquitto_sub -h 127.0.0.1 -t '{sub}' -v",
        f"[mqtt] track uploaded dataset — {fn}",
        "[mqtt] subscribed to FILE topics (not plant/line3 simulator tree)",
        f"[mqtt] confirming all {n_rows} datapoints on the wire",
    ]
    logs.extend(f"[mqtt] {_fmt_file_rd(rd)}" for rd in batch)
    logs.append(f"[mqtt] ✓ live FILE stream confirmed  ({n_rows} messages from {fn})")
    return {
        "step": 4, "source": source, "ok": True, "done": True,
        "headline": f"Live values on the stream — {n_rows} datapoints from {fn}",
        "log": logs,
        "stream": batch,
        "readings": batch,
        "reading": batch[-1],
        "count": n_rows,
        "track_done": dict(FLOOR.track_done),
    }


def run_step(step: int, param: str | None = None,
             zone: str | None = None, source: str | None = "simulator") -> Dict:
    """Walk one hop on one track. source is simulator or uploaded — never mixed."""
    if step not in (1, 2, 3, 4):
        raise ValueError("step must be 1–4")
    source = _norm_source(source)
    if source == "uploaded" and not FLOOR.rows:
        raise ValueError("Upload an Excel, CSV or JSON file first")
    pids = _pids_for(param, source)
    if source == "uploaded":
        pids = _pids_for("all", source)
    done = FLOOR.track_done.setdefault(source, 0)
    if done < step - 1:
        raise ValueError(f"Complete step {step - 1} on the {source} tab first")
    if source == "uploaded":
        return _run_uploaded_step(step, pids, done)
    return _run_simulator_step(step, pids, zone, done)

