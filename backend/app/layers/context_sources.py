"""
Layer 3 lookup catalog — uploaded Asset / MES / SAP / RDBMS tables.

Falls back to the built-in reference_data mocks when a source has not been
uploaded. Lookups stay keyed the same way Flink would key them.
"""
from __future__ import annotations
import csv
import io
import json
import mimetypes
import re
from threading import Lock
from typing import Dict, List, Optional

from ..domain import PARAMETERS, TAG_TO_PARAM
from .. import reference_data as ref
from . import context_db

KINDS = context_db.KINDS

KIND_META = {
    "asset": {
        "label": "Asset Model",
        "ic": "🗺️",
        "key": "tag",
        "hint": "tag, asset, name, unit, control_lo, control_hi, line",
        "description": "Equipment and tag master. Flink joins Gold-lake readings on tag.",
    },
    "mes": {
        "label": "MES",
        "ic": "🏭",
        "key": "asset",
        "hint": "asset, batch, product, phase, start, end, operator_shift",
        "description": "Batch and phase context. Flink joins on asset + timestamp window.",
    },
    "sap": {
        "label": "SAP",
        "ic": "🗄️",
        "key": "product",
        "hint": "product, material_no, family, grade, equipment, spec_source",
        "description": "Material master. Flink joins on product from the MES hop.",
    },
    "rdbms": {
        "label": "RDBMS / Files",
        "ic": "💾",
        "key": "tag",
        "hint": "tag, probe_calibration_date, calibration_age_days, last_maintenance, lab_note_ref",
        "description": "Lab / calibration files. Flink joins on tag.",
    },
    "other": {
        "label": "Additional source",
        "ic": "📄",
        "key": "",
        "hint": "Any tabular lookup",
        "description": "Extra dataset. Stored for reference; the Flink join uses Asset, MES, SAP and RDBMS.",
    },
}

_NAME_HINTS = {
    "asset": ("asset", "model", "equipment_master", "tag_list"),
    "mes": ("mes", "batch", "work_order", "manufacturing"),
    "sap": ("sap", "material_master", "sku_master"),
    "rdbms": ("rdbms", "lab", "calibrat", "lims", "historian"),
}
_COL_HINTS = {
    "asset": ("tag", "control_lo", "control_hi", "unit", "line", "name"),
    "mes": ("batch", "phase", "operator_shift", "asset", "start", "end"),
    "sap": ("product", "material_no", "family", "grade", "spec_source"),
    "rdbms": ("probe_calibration_date", "calibration_age_days", "last_maintenance",
              "lab_note_ref", "drift_mv"),
}


def _norm_key(s: str) -> str:
    return str(s or "").strip().lower().replace(" ", "_").replace("-", "_")


def parse_table(filename: str, data: bytes) -> List[Dict]:
    name = (filename or "").lower()
    if not data:
        raise ValueError("File is empty")
    if name.endswith(".json") or (not name and data.lstrip()[:1] in (b"{", b"[")):
        payload = json.loads(data.decode("utf-8-sig"))
        if isinstance(payload, dict):
            payload = payload.get("rows") or payload.get("data") or payload.get("records") or []
        if not isinstance(payload, list):
            raise ValueError("JSON must be an array of objects")
        return [_norm_row(r) for r in payload if isinstance(r, dict)]
    if name.endswith((".xlsx", ".xlsm")):
        return _xlsx(data)
    if name.endswith(".xls"):
        return _xls(data)
    if name.endswith(".csv") or name.endswith(".tsv") or name.endswith(".txt"):
        return _csv(data, "\t" if name.endswith(".tsv") else ",")
    raise ValueError("not a table")


def _ext(filename: str) -> str:
    name = filename or ""
    if "." not in name:
        return ""
    return "." + name.rsplit(".", 1)[-1].lower()


def _guess_mime(filename: str, data: bytes) -> str:
    if data[:4] == b"%PDF":
        return "application/pdf"
    if data[:8] == b"\x89PNG\r\n\x1a\n":
        return "image/png"
    if data[:3] == b"\xff\xd8\xff":
        return "image/jpeg"
    if data[:6] in (b"GIF87a", b"GIF89a"):
        return "image/gif"
    if data[:4] == b"RIFF" and data[8:12] == b"WEBP":
        return "image/webp"
    guessed = mimetypes.guess_type(filename or "")[0]
    return guessed or "application/octet-stream"


def _pdf_rows(data: bytes) -> List[Dict]:
    try:
        from pypdf import PdfReader
        reader = PdfReader(io.BytesIO(data))
        out = []
        for i, page in enumerate(reader.pages, 1):
            text = (page.extract_text() or "").strip()
            if text:
                out.append({"page": i, "text": text[:4000]})
        return out
    except Exception:
        return []


def _try_table(filename: str, data: bytes) -> Optional[List[Dict]]:
    ext = _ext(filename)
    table_ext = {".xlsx", ".xlsm", ".xls", ".xlsb", ".csv", ".tsv", ".txt", ".json"}
    candidates = []
    if ext in table_ext:
        candidates.append(filename)
    head = data.lstrip()[:1]
    if head in (b"{", b"["):
        candidates.append("upload.json")
    elif ext not in table_ext:
        try:
            sample = data.decode("utf-8-sig")[:4000]
            first = sample.splitlines()[0] if sample else ""
            if "," in first and len(first.split(",")) >= 2:
                candidates.append("upload.csv")
        except UnicodeDecodeError:
            pass
    seen = set()
    for name in candidates:
        if name in seen:
            continue
        seen.add(name)
        try:
            rows = parse_table(name, data)
            if rows:
                return rows
        except (ValueError, json.JSONDecodeError, UnicodeDecodeError, KeyError, TypeError):
            continue
    return None


def read_any(filename: str, data: bytes) -> Dict:
    """Accept any extension (tables, PDF, images, other binaries)."""
    if not data:
        raise ValueError("File is empty")
    ctype = _guess_mime(filename, data)
    ext = _ext(filename)
    rows = _try_table(filename, data)
    if rows:
        return {
            "rows": rows,
            "is_file": False,
            "content_type": ctype,
            "file_bytes": None,
            "size_bytes": len(data),
            "extension": ext.lstrip(".") or "table",
        }
    extra = _pdf_rows(data) if (ctype == "application/pdf" or ext == ".pdf") else []
    meta = {
        "filename": filename,
        "content_type": ctype,
        "extension": ext.lstrip(".") or "(none)",
        "size_bytes": len(data),
        "file_kind": (
            "image" if ctype.startswith("image/") else
            "pdf" if "pdf" in ctype or ext == ".pdf" else
            "file"
        ),
    }
    return {
        "rows": extra or [meta],
        "is_file": True,
        "content_type": ctype,
        "file_bytes": data,
        "size_bytes": len(data),
        "extension": meta["extension"],
    }


def _norm_row(row: Dict) -> Dict:
    out = {}
    for k, v in row.items():
        if k is None:
            continue
        key = _norm_key(str(k))
        if v is None:
            out[key] = ""
        else:
            out[key] = v if not isinstance(v, str) else v.strip()
    return out


def _xlsx(data: bytes) -> List[Dict]:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        raise ValueError("Excel sheet is empty")
    headers = [_norm_key(c) for c in rows[0]]
    out = []
    for row in rows[1:]:
        rec = {}
        empty = True
        for i, h in enumerate(headers):
            if not h:
                continue
            val = row[i] if i < len(row) else None
            if val is not None and str(val).strip() != "":
                empty = False
            rec[h] = "" if val is None else val
        if not empty:
            out.append(rec)
    return out


def _xls(data: bytes) -> List[Dict]:
    try:
        import xlrd
    except ImportError as e:
        raise ValueError("Excel .xls support is not installed (xlrd)") from e
    book = xlrd.open_workbook(file_contents=data)
    sheet = book.sheet_by_index(0)
    if sheet.nrows < 2:
        raise ValueError("Excel sheet is empty")
    headers = []
    for c in range(sheet.ncols):
        headers.append(_norm_key(str(sheet.cell_value(0, c) or "")))
    out = []
    for r in range(1, sheet.nrows):
        rec = {}
        empty = True
        for c, h in enumerate(headers):
            if not h:
                continue
            val = sheet.cell_value(r, c)
            if val not in (None, ""):
                empty = False
            rec[h] = "" if val in (None, "") else val
        if not empty:
            out.append(rec)
    return out


def _csv(data: bytes, delim: str) -> List[Dict]:
    text = data.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text), delimiter=delim)
    if not reader.fieldnames:
        raise ValueError("CSV has no header row")
    out = []
    for raw in reader:
        rec = _norm_row(raw)
        if any(str(v).strip() for v in rec.values()):
            out.append(rec)
    return out


def _floats(row: Dict, *keys) -> Optional[List[float]]:
    for k in keys:
        raw = row.get(k)
        if raw in (None, ""):
            continue
        if isinstance(raw, (list, tuple)) and len(raw) >= 2:
            return [float(raw[0]), float(raw[1])]
        s = str(raw).replace("–", "-").replace("—", "-")
        for sep in (",", ";", "-", "/"):
            if sep in s:
                parts = [p.strip() for p in s.split(sep) if p.strip()]
                if len(parts) >= 2:
                    try:
                        return [float(parts[0]), float(parts[1])]
                    except ValueError:
                        pass
    lo, hi = row.get("control_lo"), row.get("control_hi")
    if lo not in (None, "") and hi not in (None, ""):
        try:
            return [float(lo), float(hi)]
        except ValueError:
            return None
    return None


def _cell(row: Dict, *keys) -> str:
    for k in keys:
        v = row.get(k)
        if v not in (None, ""):
            return str(v).strip()
    return ""


def _as_asset(row: Dict) -> Optional[Dict]:
    tag = str(row.get("tag") or row.get("file_tag") or "").strip()
    if not tag:
        return None
    pid = TAG_TO_PARAM.get(tag) or row.get("param") or row.get("parameter")
    p = PARAMETERS.get(pid) if pid else None
    spec = _floats(row, "control", "spec", "control_band")
    if spec is None and p:
        spec = list(p.control)
    alarm = _floats(row, "alarm") or (list(p.alarm) if p else None)
    trip = _floats(row, "trip") or (list(p.trip) if p else None)
    return {
        "tag": tag,
        "asset": str(row.get("asset") or (p.asset if p else "")),
        "parameter": pid or (p.id if p else ""),
        "name": str(row.get("name") or (p.name if p else tag)),
        "unit": str(row.get("unit") or (p.unit if p else "")),
        "control": spec or [0.0, 0.0],
        "alarm": alarm,
        "trip": trip,
        "line": str(row.get("line") or ""),
    }


def _as_mes(row: Dict) -> Optional[Dict]:
    asset = _cell(row, "asset", "equipment", "asset_id", "unit")
    if not asset:
        return None
    return {
        "asset": asset,
        "batch": _cell(row, "batch", "batch_id", "batch_no", "batch_number", "batchnumber",
                       "work_order", "wo", "lot", "lot_no", "lot_id", "order", "order_id"),
        "product": _cell(row, "product", "product_name", "sku", "material_desc"),
        "phase": _cell(row, "phase", "operation", "op", "step"),
        "start": _cell(row, "start", "start_ts", "start_time", "batch_start"),
        "end": _cell(row, "end", "end_ts", "end_time", "batch_end"),
        "operator_shift": _cell(row, "operator_shift", "shift", "crew"),
    }


def _as_sap(row: Dict) -> Optional[Dict]:
    product = str(row.get("product") or "").strip()
    if not product:
        return None
    return {
        "product": product,
        "material_no": str(row.get("material_no") or row.get("material") or ""),
        "family": str(row.get("family") or ""),
        "grade": str(row.get("grade") or ""),
        "equipment": str(row.get("equipment") or row.get("asset") or ""),
        "spec_source": str(row.get("spec_source") or row.get("sop") or ""),
    }


def _as_rdbms(row: Dict) -> Optional[Dict]:
    tag = str(row.get("tag") or "").strip()
    if not tag:
        return None
    age = row.get("calibration_age_days")
    try:
        age = int(float(age)) if age not in (None, "") else None
    except (TypeError, ValueError):
        age = None
    return {
        "tag": tag,
        "probe_calibration_date": str(row.get("probe_calibration_date") or row.get("calibration_date") or ""),
        "calibration_age_days": age,
        "drift_mv": row.get("drift_mv"),
        "last_maintenance": str(row.get("last_maintenance") or ""),
        "lab_note_ref": str(row.get("lab_note_ref") or row.get("lab_note") or ""),
    }


_NORMALIZERS = {
    "asset": _as_asset,
    "mes": _as_mes,
    "sap": _as_sap,
    "rdbms": _as_rdbms,
}


def detect_kind(filename: str, rows: List[Dict]) -> str:
    """Classify a file from its name and column set."""
    name = (filename or "").lower()
    cols = set()
    if rows:
        cols = {_norm_key(k) for k in (rows[0] or {}).keys()}
    scores = {k: 0 for k in KINDS}
    for kind, needles in _NAME_HINTS.items():
        if any(n in name for n in needles):
            scores[kind] += 6
    for kind, hints in _COL_HINTS.items():
        scores[kind] += sum(1 for h in hints if h in cols)
    best = max(scores, key=scores.get)
    if scores[best] >= 2:
        return best
    return "other"


def _slot_slug(filename: str) -> str:
    stem = (filename or "source").rsplit(".", 1)[0]
    slug = re.sub(r"[^a-z0-9]+", "_", stem.lower()).strip("_")
    return slug[:48] or "source"


def _clean_rows(kind: str, rows: List[Dict]) -> tuple:
    if kind in _NORMALIZERS:
        cleaned = []
        for r in rows:
            rec = _NORMALIZERS[kind](r if isinstance(r, dict) else {})
            if rec:
                cleaned.append(rec)
        if cleaned:
            return cleaned, kind
        return [r for r in rows if isinstance(r, dict)], "other"
    return [r for r in rows if isinstance(r, dict)], "other"


def _columns_of(rows: List[Dict]) -> List[str]:
    if not rows:
        return []
    seen = []
    for k in rows[0].keys():
        if k not in seen and not str(k).startswith("_"):
            seen.append(str(k))
    return seen


def _pretty_label(filename: str) -> str:
    stem = (filename or "Dataset").rsplit(".", 1)[0]
    stem = stem.replace("_", " ").replace("-", " ").strip()
    return stem or "Additional source"


class ContextCatalog:
    def __init__(self) -> None:
        self._lock = Lock()
        self.slots: Dict[str, Dict] = {}
        self._seq = 0

    def _meta(self, kind: str, filename: str, rows: List[Dict],
              is_file: bool = False, content_type: str = "",
              size_bytes: int = 0, extension: str = "") -> Dict:
        kind = kind if kind in KIND_META else "other"
        meta = KIND_META[kind]
        cols = _columns_of(rows)
        label = meta["label"] if kind != "other" else _pretty_label(filename)
        desc = meta["description"]
        ic = meta["ic"]
        if is_file or kind == "other":
            size = size_bytes or (rows[0].get("size_bytes") if rows else 0) or 0
            ctype = content_type or ((rows[0] or {}).get("content_type") if rows else "") or ""
            ext = extension or ((rows[0] or {}).get("extension") if rows else "") or ""
            if ctype.startswith("image/"):
                ic = "🖼️"
                desc = f"Image ({ctype or ext or 'image'}, {size} bytes). Stored in PostgreSQL."
            elif "pdf" in (ctype or "") or str(ext).lower() == "pdf":
                ic = "📕"
                n = len(rows)
                desc = (
                    f"PDF document ({n} extracted page{'s' if n != 1 else ''}, {size} bytes). "
                    "Stored in PostgreSQL."
                )
            elif is_file:
                ic = "📄"
                desc = (
                    f"File {filename or ''} ({ctype or ext or 'binary'}, {size} bytes). "
                    "Stored in PostgreSQL."
                )
            elif kind == "other":
                desc = (
                    f"Additional dataset from {filename or 'upload'} "
                    f"({len(rows)} rows). Not used in the Flink join unless classified "
                    "as Asset, MES, SAP or RDBMS."
                )
        return {
            "kind": kind,
            "label": label,
            "ic": ic,
            "key": meta["key"] or (cols[0] if cols else ""),
            "hint": ", ".join(cols) if cols else meta["hint"],
            "description": desc,
            "columns": cols,
        }

    def _put(self, slot: str, kind: str, filename: str, rows: List[Dict],
             persisted: bool = False, dataset_id=None, file_bytes=None,
             content_type: str = "", is_file: bool = False,
             size_bytes: int = 0, extension: str = "") -> None:
        self._seq += 1
        prev = self.slots.get(slot) or {}
        size_bytes = size_bytes or (len(file_bytes) if file_bytes else 0)
        meta = self._meta(kind, filename, rows, is_file=is_file,
                          content_type=content_type, size_bytes=size_bytes,
                          extension=extension)
        self.slots[slot] = {
            **meta,
            "slot": slot,
            "filename": filename,
            "rows": list(rows),
            "row_count": len(rows),
            "persisted": persisted,
            "dataset_id": dataset_id,
            "loaded": bool(rows) or bool(file_bytes),
            "seq": prev.get("seq") or self._seq,
            "is_file": bool(is_file or file_bytes),
            "content_type": content_type or "",
            "file_bytes": file_bytes,
            "size_bytes": size_bytes,
            "extension": extension or _ext(filename).lstrip("."),
        }

    def _slot_for_filename(self, filename: str) -> str:
        for sid, rec in self.slots.items():
            if (rec.get("filename") or "").lower() == (filename or "").lower():
                return sid
        base = _slot_slug(filename)
        if base not in self.slots:
            return base
        n = 2
        while f"{base}_{n}" in self.slots:
            n += 1
        return f"{base}_{n}"

    def _join_rows(self, kind: str) -> List[Dict]:
        cands = [s for s in self.slots.values()
                 if s.get("kind") == kind and s.get("rows")]
        if not cands:
            return []
        cands.sort(key=lambda s: (bool(s.get("persisted")), s.get("seq", 0)))
        return cands[-1]["rows"]

    def status(self) -> Dict:
        with self._lock:
            sources = {}
            for sid, rec in self.slots.items():
                item = {
                    k: rec[k] for k in (
                        "slot", "kind", "label", "ic", "key", "hint",
                        "description", "columns", "filename", "row_count",
                        "persisted", "dataset_id", "loaded", "is_file",
                        "content_type", "size_bytes", "extension",
                    ) if k in rec
                }
                ctype = rec.get("content_type") or ""
                if rec.get("file_bytes") and ctype.startswith("image/"):
                    item["preview"] = f"/api/context/file/{sid}"
                sources[sid] = item
            return {"sources": sources, "postgres": context_db.ping()}

    def ingest_bytes(self, filename: str, data: bytes,
                     slot: Optional[str] = None) -> Dict:
        parsed = read_any(filename, data)
        rows = parsed["rows"]
        is_file = parsed["is_file"]
        if is_file:
            kind = "other"
            cleaned = rows
        else:
            kind = detect_kind(filename, rows)
            cleaned, kind = _clean_rows(kind, rows)
            if not cleaned:
                kind = "other"
                cleaned = rows
        if not cleaned and not parsed.get("file_bytes"):
            raise ValueError(f"{filename or 'File'} is empty")
        with self._lock:
            sid = slot or self._slot_for_filename(filename)
            self._put(
                sid, kind, filename, cleaned, persisted=False, dataset_id=None,
                file_bytes=parsed.get("file_bytes"),
                content_type=parsed.get("content_type") or "",
                is_file=is_file,
                size_bytes=parsed.get("size_bytes") or 0,
                extension=parsed.get("extension") or "",
            )
        return self.status()

    def ingest_file(self, filename: str, rows: List[Dict],
                    slot: Optional[str] = None) -> Dict:
        kind = detect_kind(filename, rows)
        cleaned, kind = _clean_rows(kind, rows)
        if not cleaned:
            raise ValueError(f"{filename or 'File'} has no usable rows")
        with self._lock:
            sid = slot or self._slot_for_filename(filename)
            self._put(sid, kind, filename, cleaned, persisted=False, dataset_id=None)
        return self.status()

    def load(self, kind: str, rows: List[Dict], filename: str,
             persisted: bool = False, dataset_id=None) -> Dict:
        if kind not in KINDS:
            return self.ingest_file(filename, rows)
        cleaned, detected = _clean_rows(kind, rows)
        if not cleaned or detected != kind:
            raise ValueError(
                f"{KIND_META[kind]['label']} file has no usable rows. "
                f"Columns: {KIND_META[kind]['hint']}"
            )
        with self._lock:
            self._put(kind, kind, filename, cleaned, persisted, dataset_id)
        return self.status()

    def restore_slot(self, rec: Dict) -> None:
        slot = rec.get("slot") or rec.get("kind") or _slot_slug(rec.get("filename") or "source")
        kind = rec.get("kind") or "other"
        rows = rec.get("rows") or []
        is_file = bool(rec.get("is_file") or rec.get("file_bytes"))
        if kind in _NORMALIZERS and not is_file:
            cleaned, kind = _clean_rows(kind, rows)
            rows = cleaned or rows
        with self._lock:
            self._put(
                slot, kind, rec.get("filename") or slot, rows,
                persisted=True, dataset_id=rec.get("id"),
                file_bytes=rec.get("file_bytes"),
                content_type=rec.get("content_type") or "",
                is_file=is_file,
                size_bytes=len(rec.get("file_bytes") or b"") or 0,
                extension=_ext(rec.get("filename") or "").lstrip("."),
            )

    def snapshot(self) -> Dict[str, Dict]:
        with self._lock:
            return {
                sid: {
                    "slot": sid,
                    "kind": rec.get("kind"),
                    "filename": rec.get("filename"),
                    "description": rec.get("description") or "",
                    "rows": list(rec.get("rows") or []),
                    "file_bytes": rec.get("file_bytes"),
                    "content_type": rec.get("content_type") or "",
                    "is_file": bool(rec.get("is_file") or rec.get("file_bytes")),
                }
                for sid, rec in self.slots.items()
                if rec.get("rows") or rec.get("file_bytes")
            }

    def file_payload(self, slot: str) -> Optional[Dict]:
        with self._lock:
            rec = self.slots.get(slot)
            if not rec or not rec.get("file_bytes"):
                return None
            return {
                "data": rec["file_bytes"],
                "content_type": rec.get("content_type") or "application/octet-stream",
                "filename": rec.get("filename") or slot,
            }

    def mark_persisted(self, kind: str, dataset_id, filename: str) -> None:
        """kind may be a join kind or a slot id."""
        with self._lock:
            rec = self.slots.get(kind)
            if rec is None:
                rec = next((s for s in self.slots.values() if s.get("kind") == kind), None)
            if rec is None:
                return
            rec["persisted"] = True
            rec["dataset_id"] = dataset_id
            if filename:
                rec["filename"] = filename

    def drop_slot(self, slot: str) -> Dict:
        with self._lock:
            self.slots.pop(slot, None)
        return self.status()

    def lookup_asset(self, tag: str) -> Optional[Dict]:
        with self._lock:
            for r in self._join_rows("asset"):
                if r.get("tag") == tag:
                    rec = dict(r)
                    rec["_origin"] = "upload"
                    return rec
        rec = ref.lookup_asset_model(tag)
        if rec:
            rec = dict(rec)
            rec["_origin"] = "reference"
        return rec

    def lookup_mes(self, asset: str, ts: Optional[str] = None) -> Optional[Dict]:
        with self._lock:
            candidates = [dict(r) for r in self._join_rows("mes") if r.get("asset") == asset]
        if not candidates:
            rec = ref.lookup_mes(asset, ts)
            if rec:
                rec = dict(rec)
                rec["_origin"] = "reference"
            return rec

        def _in_window(row: Dict) -> Optional[bool]:
            if not (ts and row.get("start") and row.get("end")):
                return None
            try:
                from datetime import datetime
                t = datetime.fromisoformat(str(ts).replace("Z", ""))
                start = datetime.fromisoformat(str(row["start"]).replace("Z", ""))
                end = datetime.fromisoformat(str(row["end"]).replace("Z", ""))
                return start <= t <= end
            except (ValueError, TypeError):
                return None

        timed = [r for r in candidates if _in_window(r) is True]
        pool = timed or [r for r in candidates if r.get("batch")] or candidates
        rec = dict(pool[0])
        rec["_origin"] = "upload"
        if not rec.get("batch"):
            fallback = ref.lookup_mes(asset, ts)
            if fallback and fallback.get("batch"):
                rec["batch"] = fallback["batch"]
                if not rec.get("product"):
                    rec["product"] = fallback.get("product") or ""
                rec["_batch_filled"] = "reference"
        win = _in_window(rec)
        rec["_time_match"] = (
            "in-window" if win is True else
            "outside-window" if win is False else
            "uploaded"
        )
        return rec

    def lookup_sap(self, product: str) -> Optional[Dict]:
        with self._lock:
            rec = next((dict(r) for r in self._join_rows("sap") if r.get("product") == product), None)
        if rec is not None:
            rec["_origin"] = "upload"
            return rec
        rec = ref.lookup_sap(product)
        if rec:
            rec = dict(rec)
            rec["_origin"] = "reference"
        return rec

    def lookup_rdbms(self, tag: str) -> Optional[Dict]:
        with self._lock:
            rec = next((dict(r) for r in self._join_rows("rdbms") if r.get("tag") == tag), None)
        if rec is not None:
            rec["_origin"] = "upload"
            return rec
        rec = ref.lookup_rdbms_files(tag)
        if rec:
            rec = dict(rec)
            rec["_origin"] = "reference"
        return rec


CATALOG = ContextCatalog()


def sample_csv(kind: str) -> str:
    if kind == "asset":
        lines = ["tag,asset,name,unit,control_lo,control_hi,line"]
        for p in PARAMETERS.values():
            line = {"BR-12": "Line 3", "FIL-07": "Line 3",
                    "WFI-02": "Utility", "CR-A1": "Cleanroom A"}.get(p.asset, "Line 3")
            lines.append(f"{p.tag},{p.asset},{p.name},{p.unit},{p.control[0]},{p.control[1]},{line}")
        return "\n".join(lines) + "\n"
    if kind == "mes":
        lines = ["asset,batch,product,phase,start,end,operator_shift"]
        for asset, rec in ref.MES_BATCHES.items():
            lines.append(
                f"{asset},{rec['batch']},{rec['product']},{rec['phase']},"
                f"{rec['start']},{rec['end']},{rec['operator_shift']}"
            )
        return "\n".join(lines) + "\n"
    if kind == "sap":
        lines = ["product,material_no,family,grade,equipment,spec_source"]
        for product, rec in ref.SAP_MASTER.items():
            lines.append(
                f"{product},{rec['material_no']},{rec['family']},{rec['grade']},"
                f"{rec['equipment']},{rec['spec_source']}"
            )
        return "\n".join(lines) + "\n"
    lines = ["tag,probe_calibration_date,calibration_age_days,drift_mv,last_maintenance,lab_note_ref"]
    for tag, rec in ref.RDBMS_FILES.items():
        lines.append(
            f"{tag},{rec['probe_calibration_date']},{rec['calibration_age_days']},"
            f"{rec['drift_mv']},{rec['last_maintenance']},{rec['lab_note_ref']}"
        )
    return "\n".join(lines) + "\n"
