"""
Build RAG KB — parse SOP / CAPA / OEM / Master Index / Regulatory files,
chunk, embed, index.

Extraction uses pypdf when installed; generated sample PDFs are uncompressed
text so a small fallback parser still works. Embeddings are the same TF-IDF
index SENTRA already searches (swap-in point for BGE-small + Qdrant).
"""
from __future__ import annotations
import io
import mimetypes
import re
from pathlib import Path
from threading import Lock
from typing import Dict, List, Optional, Tuple

from ..domain import PARAMETERS
from . import knowledge as kb
from . import rag_db

KINDS = rag_db.KINDS
KIND_META = {
    "capa": {
        "label": "CAPA",
        "tag": "CAPA",
        "type": "CAPA",
        "ic": "📋",
        "hint": "CAPA records — PDF, Excel, text or image",
        "description": "Corrective and preventive action records retrieved during excursions.",
    },
    "master_index": {
        "label": "Master Index",
        "tag": "INDEX",
        "type": "MASTER_INDEX",
        "ic": "📑",
        "hint": "Document master index — Excel, CSV, PDF or image",
        "description": "Master index of governed documents and their identifiers.",
    },
    "oem": {
        "label": "OEM",
        "tag": "OEM",
        "type": "OEM",
        "ic": "🛠️",
        "hint": "Equipment manuals — PDF, text or image",
        "description": "OEM troubleshooting and equipment limits for plant assets.",
    },
    "regulatory": {
        "label": "Regulatory",
        "tag": "REG",
        "type": "REGULATORY",
        "ic": "⚖️",
        "hint": "GxP / 21 CFR / annex — PDF, text or image",
        "description": "Regulatory and GxP requirements including 21 CFR Part 11.",
    },
    "sop": {
        "label": "SOP",
        "tag": "SOP",
        "type": "SOP",
        "ic": "📘",
        "hint": "Standard operating procedures — PDF, Excel, text or image",
        "description": "Standard operating procedures used when SCADA thresholds are breached.",
    },
    "other": {
        "label": "Additional source",
        "tag": "DOC",
        "type": "OTHER",
        "ic": "📄",
        "hint": "Any RAG document — PDF, Excel, image or other",
        "description": "Additional knowledge file. Indexed with the RAG pipeline when saved.",
    },
}
KIND_TYPE = {k: v["type"] for k, v in KIND_META.items()}
RAG_DIR = Path(__file__).resolve().parents[2] / "data" / "rag"

SAMPLES = [
    {
        "kind": "sop",
        "type": "SOP",
        "tag": "SOP",
        "title": "Excursion Response",
        "subtitle": "SOP-BSC-001 — all sections",
        "filename": "SOP-BSC-001_Excursion_Response.pdf",
    },
    {
        "kind": "capa",
        "type": "CAPA",
        "tag": "CAPA",
        "title": "CAPA Records",
        "subtitle": "CAPA-LOG-2024 — 5 records",
        "filename": "CAPA-LOG-2024_Records.pdf",
    },
    {
        "kind": "oem",
        "type": "OEM",
        "tag": "OEM",
        "title": "Equipment Manual",
        "subtitle": "OEM-MAN-CW3 — CW-3 / FIL-07 / HVAC",
        "filename": "OEM-MAN-CW3_Troubleshooting.pdf",
    },
]


def _wrap(text: str, width: int = 86) -> List[str]:
    lines: List[str] = []
    for para in (text or "").split("\n"):
        words = para.split()
        if not words:
            lines.append("")
            continue
        cur = words[0]
        for w in words[1:]:
            if len(cur) + 1 + len(w) <= width:
                cur += " " + w
            else:
                lines.append(cur)
                cur = w
        lines.append(cur)
    return lines


def _pdf_escape(s: str) -> str:
    return (s or "").replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def build_pdf(title: str, sections: List[Tuple[str, str]]) -> bytes:
    """Minimal uncompressed PDF so sample downloads work without reportlab."""
    rows = [title, "MOSAIC knowledge base — governed GxP content", ""]
    for heading, body in sections:
        rows.append(heading)
        rows.extend(_wrap(body, 88))
        rows.append("")
    per_page = 46
    pages = [rows[i:i + per_page] for i in range(0, max(len(rows), 1), per_page)] or [[title]]
    n = len(pages)
    font_id = 3 + 2 * n
    page_objs = []
    content_objs = []
    for i, page_lines in enumerate(pages):
        stream = "\n".join(
            ["BT", "/F1 11 Tf", "14 TL", "1 0 0 1 50 760 Tm"]
            + [f"({_pdf_escape(line[:120])}) Tj T*" for line in page_lines]
            + ["ET"]
        ).encode("latin-1", errors="replace")
        page_objs.append(
            f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] "
            f"/Resources << /Font << /F1 {font_id} 0 R >> >> "
            f"/Contents {3 + n + i} 0 R >>".encode("latin-1")
        )
        content_objs.append(
            f"<< /Length {len(stream)} >>\nstream\n".encode("latin-1") + stream + b"\nendstream"
        )
    objects = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        f"<< /Type /Pages /Kids [{' '.join(f'{3+i} 0 R' for i in range(n))}] /Count {n} >>".encode("latin-1"),
        *page_objs,
        *content_objs,
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
    ]
    buf = io.BytesIO()
    buf.write(b"%PDF-1.4\n")
    offsets = [0]
    for i, obj in enumerate(objects, start=1):
        offsets.append(buf.tell())
        buf.write(f"{i} 0 obj\n".encode("latin-1"))
        buf.write(obj)
        buf.write(b"\nendobj\n")
    xref = buf.tell()
    buf.write(f"xref\n0 {len(objects)+1}\n".encode("latin-1"))
    buf.write(b"0000000000 65535 f \n")
    for off in offsets[1:]:
        buf.write(f"{off:010d} 00000 n \n".encode("latin-1"))
    buf.write(
        f"trailer << /Size {len(objects)+1} /Root 1 0 R >>\nstartxref\n{xref}\n%%EOF\n".encode("latin-1")
    )
    return buf.getvalue()


def sample_sections(kind: str) -> List[Tuple[str, str]]:
    kind = (kind or "").lower()
    want = {"sop": "SOP", "capa": "CAPA", "oem": "OEM"}.get(kind)
    if not want:
        return []
    out = []
    for rec in kb.SEED:
        if rec.get("type") == want:
            out.append((rec.get("ref") or rec["id"], rec["text"]))
    return out


def sample_pdf(kind: str) -> Tuple[bytes, str]:
    meta = next((s for s in SAMPLES if s["kind"] == kind.lower()), None)
    if not meta:
        raise ValueError("kind must be sop, capa or oem")
    sections = sample_sections(kind)
    data = build_pdf(f"{meta['tag']} — {meta['title']}", sections)
    return data, meta["filename"]


def detect_type(filename: str, override: Optional[str] = None) -> str:
    if override and override.lower() not in ("", "auto"):
        key = override.lower().replace(" ", "_").replace("-", "_")
        if key in KIND_TYPE:
            return KIND_TYPE[key]
        if key == "other":
            return "OTHER"
        return override.upper()
    name = (filename or "").lower()
    if "capa" in name:
        return "CAPA"
    if "master" in name or "index" in name:
        return "MASTER_INDEX"
    if "regulat" in name or "cfr" in name or "annex" in name or "gxp" in name:
        return "REGULATORY"
    if "oem" in name or "manual" in name:
        return "OEM"
    if "sop" in name:
        return "SOP"
    return "OTHER"


def detect_kind(filename: str) -> str:
    name = (filename or "").lower()
    if "capa" in name:
        return "capa"
    if "master" in name or "index" in name:
        return "master_index"
    if "regulat" in name or "cfr" in name or "annex" in name or "gxp" in name:
        return "regulatory"
    if "oem" in name or "manual" in name:
        return "oem"
    if "sop" in name:
        return "sop"
    return "other"


def _slot_slug(filename: str) -> str:
    stem = (filename or "source").rsplit(".", 1)[0]
    slug = re.sub(r"[^a-z0-9]+", "_", stem.lower()).strip("_")
    return slug[:48] or "source"


def _guess_mime(filename: str, data: bytes) -> str:
    if data[:4] == b"%PDF":
        return "application/pdf"
    if data[:8] == b"\x89PNG\r\n\x1a\n":
        return "image/png"
    if data[:3] == b"\xff\xd8\xff":
        return "image/jpeg"
    if data[:6] in (b"GIF87a", b"GIF89a"):
        return "image/gif"
    if data[:4] == b"RIFF" and data[8:12] == b"WEBP":
        return "image/webp"
    return mimetypes.guess_type(filename or "")[0] or "application/octet-stream"


def _excel_columns(data: bytes, xls: bool = False) -> List[str]:
    try:
        if xls:
            import xlrd
            book = xlrd.open_workbook(file_contents=data)
            sheet = book.sheet_by_index(0)
            return [str(sheet.cell_value(0, c) or "").strip() for c in range(sheet.ncols) if str(sheet.cell_value(0, c) or "").strip()]
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        ws = wb.active
        row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        wb.close()
        if not row:
            return []
        return [str(c).strip() for c in row if c is not None and str(c).strip()]
    except Exception:
        return []


def inspect_file(filename: str, data: bytes) -> Dict:
    name = (filename or "").lower()
    ctype = _guess_mime(filename, data or b"")
    ext = name.rsplit(".", 1)[-1] if "." in name else ""
    cols: List[str] = []
    if name.endswith((".xlsx", ".xlsm")):
        cols = _excel_columns(data, xls=False)
    elif name.endswith(".xls"):
        cols = _excel_columns(data, xls=True)
    elif name.endswith((".csv", ".tsv", ".txt")):
        try:
            first = (data or b"").decode("utf-8-sig", errors="replace").splitlines()[0]
            delim = "\t" if name.endswith(".tsv") else ","
            cols = [c.strip() for c in first.split(delim) if c.strip()]
        except Exception:
            cols = []
    elif "pdf" in ctype or name.endswith(".pdf"):
        cols = ["page", "text"]
    elif ctype.startswith("image/"):
        cols = ["filename", "content_type", "size_bytes"]
    else:
        cols = ["filename", "content_type", "size_bytes"]
    return {
        "content_type": ctype,
        "extension": ext,
        "columns": cols,
        "is_image": ctype.startswith("image/"),
        "is_pdf": "pdf" in ctype or name.endswith(".pdf"),
        "size_bytes": len(data or b""),
    }


def _safe_name(filename: str) -> str:
    base = Path(filename or "document").name
    base = re.sub(r"[^A-Za-z0-9._-]+", "_", base)
    return (base[:120] or "document")


def _write_disk(kind: str, filename: str, data: bytes, slot: Optional[str] = None) -> Path:
    folder = RAG_DIR / (slot or kind or "other")
    folder.mkdir(parents=True, exist_ok=True)
    for p in folder.iterdir():
        if p.is_file():
            p.unlink()
    path = folder / _safe_name(filename)
    path.write_bytes(data or b"")
    return path


def _load_disk() -> Dict[str, Dict]:
    out: Dict[str, Dict] = {}
    if not RAG_DIR.exists():
        return out
    for folder in RAG_DIR.iterdir():
        if not folder.is_dir():
            continue
        files = [p for p in folder.iterdir() if p.is_file()]
        if not files:
            continue
        path = sorted(files, key=lambda p: p.stat().st_mtime, reverse=True)[0]
        slot = folder.name
        kind = slot if slot in KINDS else "other"
        out[slot] = {
            "slot": slot,
            "kind": kind,
            "filename": path.name,
            "data": path.read_bytes(),
            "byte_count": path.stat().st_size,
            "persisted": True,
            "id": None,
        }
    return out


def _guess_param(text: str) -> Optional[str]:
    low = (text or "").lower()
    hits = []
    for p in PARAMETERS.values():
        score = 0
        if p.short.lower() in low or p.name.lower() in low:
            score += 2
        if p.tag.lower() in low:
            score += 2
        if p.id in low.split():
            score += 1
        if score:
            hits.append((score, p.id))
    hits.sort(reverse=True)
    return hits[0][1] if hits else None


def _extract_xlsx(data: bytes) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    lines: List[str] = []
    try:
        for ws in wb.worksheets:
            lines.append(f"# {ws.title}")
            for row in ws.iter_rows(values_only=True):
                cells = [
                    str(c).strip()
                    for c in row
                    if c is not None and str(c).strip() != ""
                ]
                if cells:
                    lines.append(" | ".join(cells))
    finally:
        wb.close()
    return "\n".join(lines).strip()


def _extract_xls(data: bytes) -> str:
    import xlrd
    book = xlrd.open_workbook(file_contents=data)
    lines: List[str] = []
    for i in range(book.nsheets):
        sheet = book.sheet_by_index(i)
        lines.append(f"# {sheet.name}")
        for r in range(sheet.nrows):
            cells = []
            for c in range(sheet.ncols):
                val = sheet.cell_value(r, c)
                if val not in (None, ""):
                    cells.append(str(val).strip())
            if cells:
                lines.append(" | ".join(cells))
    return "\n".join(lines).strip()


def extract_text(filename: str, data: bytes) -> str:
    name = (filename or "").lower()
    if name.endswith((".txt", ".md", ".csv", ".tsv")):
        return data.decode("utf-8", errors="replace")
    if name.endswith((".xlsx", ".xlsm")):
        text = _extract_xlsx(data)
        if text:
            return text
        raise ValueError("Excel sheet is empty")
    if name.endswith(".xls"):
        text = _extract_xls(data)
        if text:
            return text
        raise ValueError("Excel sheet is empty")
    try:
        from pypdf import PdfReader
        reader = PdfReader(io.BytesIO(data))
        parts = []
        for page in reader.pages:
            parts.append(page.extract_text() or "")
        text = "\n".join(parts).strip()
        if text:
            return text
    except Exception:
        pass
    ctype = _guess_mime(filename, data or b"")
    if ctype.startswith("image/") or name.endswith((".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp", ".tif", ".tiff", ".svg")):
        return (
            f"[image] {filename}\n"
            f"content_type={ctype}\n"
            f"size_bytes={len(data or b'')}"
        )
    raw = data.decode("latin-1", errors="replace")
    found = re.findall(r"\((?:\\.|[^\\)]){3,}\)(?:\s*Tj|\s*')", raw)
    if found:
        cleaned = []
        for s in found:
            s = s.rsplit(")", 1)[0][1:]
            s = s.replace("\\n", "\n").replace("\\(", "(").replace("\\)", ")").replace("\\\\", "\\")
            cleaned.append(s)
        return "\n".join(cleaned)
    return (
        f"[file] {filename}\n"
        f"content_type={ctype}\n"
        f"size_bytes={len(data or b'')}"
    )


def chunk_text(text: str, size: int = 420) -> List[str]:
    paras = [p.strip() for p in re.split(r"\n\s*\n", text or "") if p.strip()]
    if not paras:
        paras = [text.strip()] if (text or "").strip() else []
    chunks: List[str] = []
    buf = ""
    for p in paras:
        if buf and len(buf) + 1 + len(p) > size:
            chunks.append(buf)
            buf = p
        else:
            buf = f"{buf}\n\n{p}".strip() if buf else p
    if buf:
        chunks.append(buf)
    out: List[str] = []
    for ch in chunks:
        if len(ch) <= size * 2:
            out.append(ch)
            continue
        for i in range(0, len(ch), size):
            piece = ch[i:i + size].strip()
            if piece:
                out.append(piece)
    return out


def ingest_file(
    filename: str,
    data: bytes,
    doc_type: str = "auto",
    rag_kind: Optional[str] = None,
    pipeline: bool = False,
) -> Dict:
    steps: List[Dict] = []
    size = len(data or b"")
    slot = (rag_kind or "").lower().replace(" ", "_") if rag_kind else None
    if slot in ("", "auto", "none"):
        slot = None
    if pipeline:
        steps.append({
            "name": "Load saved file",
            "detail": f"loaded {filename} ({size} bytes) from backend",
            "ok": True,
        })
    else:
        steps.append({
            "name": "Upload",
            "detail": f"received {filename} ({size} bytes)",
            "ok": True,
        })
    kind = detect_type(filename, slot or doc_type)
    text = extract_text(filename, data)
    steps.append({
        "name": "Extract text",
        "detail": f"extracted {len(text)} characters · type {kind}",
        "ok": True,
    })
    pieces = chunk_text(text)
    if not pieces:
        pieces = [text.strip()] if (text or "").strip() else [f"[file] {filename}"]
    steps.append({
        "name": "Chunk",
        "detail": f"split into {len(pieces)} chunk(s)",
        "ok": True,
    })
    if slot:
        kb.remove_rag_kind(slot)
    chunks = []
    stem = re.sub(r"[^A-Za-z0-9._-]+", "_", filename.rsplit(".", 1)[0])[:40]
    prefix = (slot or "up").replace("_", "")[:8].upper()
    for i, piece in enumerate(pieces, start=1):
        chunks.append({
            "id": f"UP-{prefix}-{stem}-{i:03d}",
            "type": kind,
            "ref": f"{stem} §{i}",
            "param": _guess_param(piece),
            "text": piece,
            "source": "upload",
            "filename": filename,
            "rag_kind": slot,
        })
    steps.append({
        "name": "Embed",
        "detail": f"embedded {len(chunks)} chunk(s) (TF-IDF · BGE-small swap-in)",
        "ok": True,
    })
    added = kb.add_chunks(chunks)
    st = kb.stats()
    steps.append({
        "name": "Index",
        "detail": f"added {added} to hybrid index · {st['chunks']} chunks total",
        "ok": True,
    })
    return {
        "filename": filename,
        "type": kind,
        "rag_kind": slot,
        "bytes": size,
        "chars": len(text),
        "chunks": len(chunks),
        "added": added,
        "index_chunks": st["chunks"],
        "steps": steps,
    }


class RagCatalog:
    def __init__(self) -> None:
        self._lock = Lock()
        self.slots: Dict[str, Dict] = {}
        self._seq = 0

    def _slot_for_filename(self, filename: str) -> str:
        for sid, rec in self.slots.items():
            if (rec.get("filename") or "").lower() == (filename or "").lower():
                return sid
        base = _slot_slug(filename)
        if base not in self.slots:
            return base
        n = 2
        while f"{base}_{n}" in self.slots:
            n += 1
        return f"{base}_{n}"

    def _describe(self, kind: str, filename: str, info: Dict) -> str:
        meta = KIND_META.get(kind) or KIND_META["other"]
        if kind in KIND_META and kind != "other":
            return meta.get("description") or meta["hint"]
        ctype = info.get("content_type") or ""
        size = info.get("size_bytes") or 0
        if info.get("is_image"):
            return f"Image file ({ctype}, {size} bytes). Stored for RAG; caption is indexed."
        if info.get("is_pdf"):
            return f"PDF document ({size} bytes). Text is extracted, chunked and indexed."
        return f"File {filename} ({ctype or info.get('extension') or 'binary'}, {size} bytes)."

    def _put(self, slot: str, kind: str, filename: str, data: bytes,
             persisted: bool = False, dataset_id=None,
             content_type: str = "", columns: Optional[List[str]] = None,
             description: str = "") -> None:
        self._seq += 1
        prev = self.slots.get(slot) or {}
        info = inspect_file(filename, data)
        kind = kind if kind in KIND_META else "other"
        meta = KIND_META[kind]
        label = meta["label"] if kind != "other" else (
            (filename or "Dataset").rsplit(".", 1)[0].replace("_", " ").replace("-", " ")
        )
        ic = meta["ic"]
        if info.get("is_image"):
            ic = "🖼️"
        elif info.get("is_pdf") and kind == "other":
            ic = "📕"
        cols = columns if columns is not None else info.get("columns") or []
        desc = description or self._describe(kind, filename, info)
        self.slots[slot] = {
            "slot": slot,
            "kind": kind,
            "label": label,
            "tag": meta["tag"],
            "type": meta["type"],
            "ic": ic,
            "hint": meta["hint"],
            "description": desc,
            "columns": cols,
            "filename": filename or f"{kind}.bin",
            "data": data or b"",
            "bytes": len(data or b""),
            "content_type": content_type or info.get("content_type") or "",
            "extension": info.get("extension") or "",
            "is_image": bool(info.get("is_image")),
            "persisted": persisted,
            "dataset_id": dataset_id,
            "loaded": bool(data),
            "seq": prev.get("seq") or self._seq,
        }

    def status(self) -> Dict:
        with self._lock:
            sources = {}
            for sid, rec in self.slots.items():
                item = {k: rec[k] for k in (
                    "slot", "kind", "label", "tag", "type", "ic", "hint",
                    "description", "columns", "filename", "bytes",
                    "content_type", "extension", "is_image",
                    "persisted", "dataset_id", "loaded",
                ) if k in rec}
                if rec.get("is_image") and rec.get("data"):
                    item["preview"] = f"/api/rag/file/{sid}"
                sources[sid] = item
            saved_count = sum(1 for s in sources.values() if s.get("persisted"))
            loaded_count = sum(1 for s in sources.values() if s.get("loaded"))
        return {
            "sources": sources,
            "saved_count": saved_count,
            "loaded_count": loaded_count,
            "pipeline_ready": saved_count > 0,
            "postgres": rag_db.ping(),
        }

    def ingest_bytes(self, filename: str, data: bytes,
                     slot: Optional[str] = None) -> Dict:
        payload = data or b""
        if not payload:
            raise ValueError("File is empty")
        kind = detect_kind(filename)
        info = inspect_file(filename, payload)
        with self._lock:
            sid = slot or self._slot_for_filename(filename)
            self._put(sid, kind, filename, payload, persisted=False, dataset_id=None,
                      content_type=info.get("content_type") or "",
                      columns=info.get("columns") or [])
        return self.status()

    def receive(self, kind: str, filename: str, data: bytes,
                persisted: bool = False, dataset_id=None) -> Dict:
        payload = data or b""
        if not payload:
            raise ValueError("File is empty")
        kind = (kind or "other").lower()
        if kind not in KIND_META:
            kind = detect_kind(filename)
        info = inspect_file(filename, payload)
        with self._lock:
            sid = kind if kind in KINDS else self._slot_for_filename(filename)
            self._put(sid, kind, filename or f"{kind}.bin", payload,
                      persisted=persisted, dataset_id=dataset_id,
                      content_type=info.get("content_type") or "",
                      columns=info.get("columns") or [])
        return self.status()

    def restore_slot(self, rec: Dict) -> None:
        slot = rec.get("slot") or rec.get("kind") or _slot_slug(rec.get("filename") or "source")
        kind = rec.get("kind") or detect_kind(rec.get("filename") or "")
        data = rec.get("data") or b""
        info = inspect_file(rec.get("filename") or "", data)
        with self._lock:
            self._put(
                slot, kind, rec.get("filename") or slot, data,
                persisted=True, dataset_id=rec.get("id"),
                content_type=rec.get("content_type") or info.get("content_type") or "",
                columns=info.get("columns") or [],
                description=rec.get("description") or "",
            )

    def snapshot(self) -> Dict[str, Dict]:
        with self._lock:
            return {
                sid: {
                    "slot": sid,
                    "kind": rec.get("kind"),
                    "filename": rec.get("filename"),
                    "description": rec.get("description") or "",
                    "content_type": rec.get("content_type") or "",
                    "data": rec.get("data") or b"",
                    "bytes": rec.get("bytes") or 0,
                    "persisted": rec.get("persisted"),
                    "dataset_id": rec.get("dataset_id"),
                }
                for sid, rec in self.slots.items()
                if rec.get("data")
            }

    def mark_persisted(self, kind: str, dataset_id, filename: str) -> None:
        with self._lock:
            rec = self.slots.get(kind)
            if rec is None:
                rec = next((s for s in self.slots.values() if s.get("kind") == kind), None)
            if rec is None:
                return
            rec["persisted"] = True
            rec["dataset_id"] = dataset_id
            if filename:
                rec["filename"] = filename

    def drop_slot(self, slot: str) -> Dict:
        with self._lock:
            self.slots.pop(slot, None)
        return self.status()

    def file_payload(self, slot: str) -> Optional[Dict]:
        with self._lock:
            rec = self.slots.get(slot)
            if not rec or not rec.get("data"):
                return None
            return {
                "data": rec["data"],
                "content_type": rec.get("content_type") or "application/octet-stream",
                "filename": rec.get("filename") or slot,
            }

    def saved_files(self) -> Dict[str, Dict]:
        with self._lock:
            return {
                sid: {
                    "slot": sid,
                    "kind": rec.get("kind"),
                    "filename": rec["filename"],
                    "data": rec["data"],
                    "bytes": rec["bytes"],
                }
                for sid, rec in self.slots.items()
                if rec and rec.get("data") and rec.get("persisted")
            }


CATALOG = RagCatalog()


def save_all() -> Dict:
    snap = CATALOG.snapshot()
    if not snap:
        raise ValueError("Upload at least one document first")
    pg = rag_db.ping()
    saved = []
    pg_ok = bool(pg.get("ok"))
    pg_error = None
    for slot, rec in snap.items():
        filename = rec.get("filename") or f"{slot}.bin"
        data = rec.get("data") or b""
        kind = rec.get("kind") or "other"
        _write_disk(kind, filename, data, slot=slot)
        dataset_id = rec.get("dataset_id")
        if pg_ok:
            try:
                row = rag_db.save_slot(
                    slot, kind, filename, data,
                    rec.get("description") or "", rec.get("content_type") or "",
                )
                dataset_id = row.get("id")
                saved.append(row)
            except Exception as e:
                pg_ok = False
                pg_error = str(e)
                saved.append({
                    "slot": slot, "kind": kind, "filename": filename,
                    "byte_count": len(data), "backend": "disk",
                })
        else:
            saved.append({
                "slot": slot, "kind": kind, "filename": filename,
                "byte_count": len(data), "backend": "disk",
            })
        CATALOG.mark_persisted(slot, dataset_id, filename)
    st = status()
    st["saved"] = saved
    st["postgres"] = pg
    if pg_ok:
        st["message"] = f"Saved {len(saved)} document(s) to PostgreSQL and the backend store"
    else:
        st["message"] = (
            f"Saved {len(saved)} document(s) to the backend store"
            + (f" (PostgreSQL unavailable: {pg_error or pg.get('error')})" if (pg_error or not pg.get("ok")) else "")
        )
    return st


def save_one(slot: str) -> Dict:
    snap = CATALOG.snapshot()
    rec = snap.get(slot)
    if not rec or not rec.get("data"):
        raise ValueError("Upload this document first, then Save")
    pg = rag_db.ping()
    filename = rec.get("filename") or f"{slot}.bin"
    data = rec.get("data") or b""
    kind = rec.get("kind") or "other"
    _write_disk(kind, filename, data, slot=slot)
    dataset_id = rec.get("dataset_id")
    saved = None
    if pg.get("ok"):
        saved = rag_db.save_slot(
            slot, kind, filename, data,
            rec.get("description") or "", rec.get("content_type") or "",
        )
        dataset_id = saved.get("id")
    CATALOG.mark_persisted(slot, dataset_id, filename)
    st = status()
    st["saved"] = saved
    st["postgres"] = pg
    st["message"] = (
        f"Saved {filename} to PostgreSQL"
        if pg.get("ok") else
        f"Saved {filename} to the backend store (PostgreSQL unavailable)"
    )
    return st


def delete_one(slot: str) -> Dict:
    pg = rag_db.ping()
    if not pg.get("ok"):
        raise ConnectionError(
            "PostgreSQL is not reachable. Start it with: docker compose up -d postgres-floor"
        )
    deleted = rag_db.delete_slot(slot)
    folder = RAG_DIR / slot
    if folder.is_dir():
        for p in folder.iterdir():
            if p.is_file():
                p.unlink()
    kb.remove_rag_kind(slot)
    st = CATALOG.drop_slot(slot)
    st.update(kb.stats())
    st["deleted"] = deleted
    st["postgres"] = pg
    st["message"] = "Document deleted from PostgreSQL. Upload the file again to restore it."
    return st


def run_pipeline() -> Dict:
    files = CATALOG.saved_files()
    if not files:
        raise ValueError("Save at least one document first, then start the RAG pipeline")
    jobs = []
    errors = []
    for slot, rec in files.items():
        try:
            jobs.append(ingest_file(
                rec["filename"], rec["data"],
                doc_type=rec.get("kind") or "auto",
                rag_kind=slot, pipeline=True,
            ))
        except Exception as e:
            errors.append({"filename": rec["filename"], "kind": rec.get("kind"), "error": str(e)})
    st = status()
    st["jobs"] = jobs
    st["errors"] = errors
    st["message"] = (
        f"Pipeline finished · {len(jobs)} document(s) indexed"
        + (f" · {len(errors)} failed" if errors else "")
    )
    return st


def restore() -> Dict:
    """Reload saved RAG files from PostgreSQL (then disk) and re-index them."""
    loaded: Dict[str, Dict] = {}
    try:
        for rec in rag_db.load_slots():
            loaded[rec["slot"]] = rec
    except Exception:
        pass
    disk = _load_disk()
    for slot, rec in disk.items():
        if slot not in loaded:
            loaded[slot] = rec
    for rec in loaded.values():
        try:
            CATALOG.restore_slot(rec)
        except (ValueError, TypeError, KeyError):
            continue
    if CATALOG.saved_files():
        try:
            return run_pipeline()
        except ValueError:
            pass
    return status()


def status() -> Dict:
    st = kb.stats()
    st["samples"] = SAMPLES
    cat = CATALOG.status()
    st.update(cat)
    return st
